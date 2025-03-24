#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script moderno para reorganizar dados de projetos de CSV para Excel.
Utiliza abordagens eficientes do pandas sem usar loops for explícitos.
"""

import pandas as pd
import os
import re
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def reorganizar_projetos(arquivo_csv, arquivo_saida='Projetos_Reorganizados.xlsx', encoding='cp1252'):
    """
    Reorganiza os dados de projetos do arquivo CSV para um formato tabular no Excel.
    Utiliza métodos vetorizados do pandas para maior eficiência.
    
    Args:
        arquivo_csv: Caminho para o arquivo CSV de entrada
        arquivo_saida: Caminho para o arquivo Excel de saída
        encoding: Codificação do arquivo CSV
    """
    print(f"Processando arquivo: {arquivo_csv}")
    
    # Ler o arquivo CSV
    try:
        df = pd.read_csv(arquivo_csv, encoding=encoding)
        print(f"Arquivo lido com sucesso. Colunas encontradas: {len(df.columns)}")
    except Exception as e:
        print(f"Erro ao ler o arquivo CSV: {str(e)}")
        return False
    
    # Identificar colunas de projetos por padrões
    colunas_por_tipo = identificar_colunas_projeto(df)
    
    # Lista de todos os pares de colunas (nome, status, versão, autor)
    pares_colunas = []
    for i in range(max(len(grupo) for grupo in colunas_por_tipo.values())):
        par = {}
        for tipo, colunas in colunas_por_tipo.items():
            if i < len(colunas):
                par[tipo] = colunas[i]
        if par:
            pares_colunas.append(par)
    
    # Criar um DataFrame vazio para armazenar projetos reorganizados
    projetos_df = pd.DataFrame(columns=[
        'Nome do Projeto', 'Status', 'Versão', 'Autor', 
        'Email Respondente', 'Nome Respondente'
    ])
    
    # Para cada conjunto de colunas relacionadas, extrair os projetos
    for par in pares_colunas:
        if 'nome' not in par:
            continue
            
        # Selecionar apenas as colunas relevantes e os respondentes
        colunas_selecionadas = [col for col in par.values() if col in df.columns]
        colunas_selecionadas.extend(['Email', 'Nome'])
        
        # Criar um DataFrame temporário com as colunas relevantes
        temp_df = df[colunas_selecionadas].copy()
        
        # Filtrar linhas onde o nome do projeto não é nulo
        temp_df = temp_df[temp_df[par['nome']].notna()]
        
        if len(temp_df) == 0:
            continue
            
        # Renomear colunas para o formato padrão
        mapeamento_colunas = {
            par.get('nome', ''): 'Nome do Projeto',
            par.get('status', ''): 'Status',
            par.get('versao', ''): 'Versão',
            par.get('autor', ''): 'Autor',
            'Email': 'Email Respondente',
            'Nome': 'Nome Respondente'
        }
        # Remover chaves vazias ou valores que não existem no DataFrame
        mapeamento_colunas = {k: v for k, v in mapeamento_colunas.items() 
                              if k and k in temp_df.columns}
        
        temp_df = temp_df.rename(columns=mapeamento_colunas)
        
        # Selecionar apenas as colunas padrão que existem
        colunas_existentes = [col for col in projetos_df.columns if col in temp_df.columns]
        temp_df = temp_df[colunas_existentes]
        
        # Concatenar com o DataFrame principal
        projetos_df = pd.concat([projetos_df, temp_df], ignore_index=True)
    
    print(f"Dados processados. Encontrados {len(projetos_df)} projetos.")
    
    # Garantir diretório de saída
    diretorio_saida = os.path.dirname(arquivo_saida)
    if diretorio_saida and not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Exportar para Excel
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        # Tabela principal
        projetos_df.to_excel(writer, sheet_name='Projetos', index=False)
        
        # Resumo por Status (usando pivot_table diretamente)
        if 'Status' in projetos_df.columns and 'Autor' in projetos_df.columns:
            pivot_status = pd.pivot_table(
                projetos_df,
                values='Nome do Projeto',
                index=['Autor'],
                columns=['Status'],
                aggfunc='count',
                fill_value=0
            )
            pivot_status.to_excel(writer, sheet_name='Resumo por Status')
        
        # Resumo por Respondente (usando groupby e size)
        if 'Nome Respondente' in projetos_df.columns:
            resumo_respondente = (projetos_df
                                  .groupby('Nome Respondente', as_index=False)
                                  .size()
                                  .rename(columns={'size': 'Total de Projetos'}))
            resumo_respondente.to_excel(writer, sheet_name='Resumo por Respondente', index=False)
    
    # Adicionar formatação ao Excel
    formatar_excel(arquivo_saida)
    
    print(f"Dados exportados com sucesso para: {arquivo_saida}")
    return True


def identificar_colunas_projeto(df):
    """
    Identifica e agrupa as colunas relacionadas a projetos no DataFrame.
    Utiliza expressões regulares para identificar padrões de colunas.
    
    Args:
        df: DataFrame pandas com os dados
        
    Returns:
        Dicionário com colunas agrupadas por tipo
    """
    # Padrões regex para identificar colunas
    padroes = {
        'nome': r'^Nome do Projeto(\d*)$',
        'status': r'^Status do Projeto.Meu projeto está:(\d*)$',
        'versao': r'^Versão do Projeto(\d*)$',
        'autor': r'^Autor \(Responsável pelo Projeto\)(\d*)$',
        'continuar': r'^Deseja adicionar outro projeto \?(\d*)$'
    }
    
    # Usar compreensões de dicionário para criar a estrutura
    colunas_por_tipo = {
        tipo: [col for col in df.columns if re.match(padrao, col)]
        for tipo, padrao in padroes.items()
    }
    
    # Ordenar cada grupo de colunas usando uma função lambda
    for tipo, padrao in padroes.items():
        if colunas_por_tipo[tipo]:
            # Extrai o número do sufixo para ordenar corretamente
            colunas_por_tipo[tipo].sort(
                key=lambda x: int(re.match(padrao, x).group(1)) if re.match(padrao, x).group(1) else 0
            )
    
    return colunas_por_tipo


def formatar_excel(arquivo_excel):
    """
    Adiciona formatação ao arquivo Excel.
    
    Args:
        arquivo_excel: Caminho para o arquivo Excel
    """
    try:
        wb = load_workbook(arquivo_excel)
        
        # Cores para formatação
        header_color = "DDEBF7"
        status_colors = {
            "Concluído": "C6EFCE",
            "Em andamento": "FFEB9C",
            "Não iniciado": "F8CBAD",
            "Cancelado": "F2F2F2"
        }
        
        # Formatar planilhas usando a função apply_sheet_formatting
        for sheet_name in ['Projetos', 'Resumo por Status', 'Resumo por Respondente']:
            if sheet_name in wb.sheetnames:
                apply_sheet_formatting(wb[sheet_name], header_color)
        
        # Salvar o arquivo
        wb.save(arquivo_excel)
    except Exception as e:
        print(f"Aviso: Não foi possível aplicar formatação avançada: {str(e)}")


def apply_sheet_formatting(worksheet, header_color):
    """
    Aplica formatação a uma planilha específica.
    
    Args:
        worksheet: Objeto de planilha do openpyxl
        header_color: Cor para o cabeçalho
    """
    # Formatar cabeçalhos
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Ajustar largura das colunas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Função principal do script."""
    parser = argparse.ArgumentParser(description='Reorganiza dados de projetos de CSV para Excel.')
    parser.add_argument('arquivo_csv', help='Caminho para o arquivo CSV de entrada')
    parser.add_argument('--saida', '-s', help='Caminho para o arquivo Excel de saída',
                      default='Projetos_Reorganizados.xlsx')
    parser.add_argument('--encoding', '-e', help='Encoding do arquivo CSV',
                      default='cp1252')
    
    args = parser.parse_args()
    
    # Verificar se o arquivo existe
    if not os.path.exists(args.arquivo_csv):
        print(f"Erro: Arquivo '{args.arquivo_csv}' não encontrado.")
        return 1
    
    # Reorganizar os dados
    resultado = reorganizar_projetos(
        args.arquivo_csv,
        args.saida,
        args.encoding
    )
    
    return 0 if resultado else 1


if __name__ == "__main__":
    main()